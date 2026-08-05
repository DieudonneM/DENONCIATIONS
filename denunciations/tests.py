import importlib
import os
from io import BytesIO
from unittest import mock

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

import denunciations_app.settings as settings_module

from api.serializers import PieceJointeSerializer
from core.models import Employeur, Province
from users.models import User
from .models import Incident, PieceJointe


@override_settings(ALLOWED_HOSTS=['testserver'])
class LegacyRouteRedirectTests(TestCase):
    def test_legacy_incidents_url_redirects_to_core_namespace(self):
        response = self.client.get(reverse('denunciations:incidents_list'))

        self.assertRedirects(
            response,
            reverse('core:incidents_list'),
            status_code=307,
            target_status_code=200,
            fetch_redirect_response=False,
        )

    def test_legacy_form_post_preserves_request_method(self):
        response = self.client.post(reverse('denunciations:incident_form'))

        self.assertEqual(response.status_code, 307)
        self.assertEqual(response['Location'], reverse('core:incident_form'))


@override_settings(ALLOWED_HOSTS=['testserver'])
class IncidentExportAuthorizationTests(TestCase):
    def test_worker_export_excludes_incidents_owned_by_other_workers(self):
        province = Province.objects.create(nom='Kinshasa', code='KIN')
        employeur = Employeur.objects.create(
            nom='Entreprise privée',
            secteur='services',
            province=province,
        )
        owner = User.objects.create_user(
            username='owner',
            email='owner@example.test',
            password='safe-password',
            role='travailleur',
        )
        other_worker = User.objects.create_user(
            username='other-worker',
            email='other@example.test',
            password='safe-password',
            role='travailleur',
        )
        private_incident = Incident.objects.create(
            code_suivi='RDC2026PRIVATE',
            travailleur=owner,
            employeur=employeur,
            province=province,
            ville='Kinshasa',
            type_incident='salaire',
            description='Incident confidentiel appartenant à un autre travailleur.',
        )

        self.client.force_login(other_worker)
        response = self.client.get(reverse('core:export_incidents_xlsx'))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        values = [
            value
            for row in workbook.active.iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        self.assertNotIn(private_incident.code_suivi, values)
        self.assertNotIn(private_incident.description, values)


class PieceJointeSerializerTests(TestCase):
    def test_serializer_exposes_attachment_url_for_mobile_clients(self):
        province = Province.objects.create(nom='Kinshasa', code='KIN')
        employeur = Employeur.objects.create(
            nom='Entreprise privée',
            secteur='services',
            province=province,
        )
        incident = Incident.objects.create(
            employeur=employeur,
            province=province,
            ville='Kinshasa',
            type_incident='salaire',
            description='Description suffisante pour tester la sérialisation des pièces jointes.',
            est_anonyme=True,
        )
        attachment = PieceJointe.objects.create(
            incident=incident,
            fichier=SimpleUploadedFile('preuve.pdf', b'abc123', content_type='application/pdf'),
            nom_original='preuve.pdf',
            type_fichier='application/pdf',
            taille_fichier=6,
        )

        data = PieceJointeSerializer(attachment).data

        self.assertIn('fichier_url', data)
        self.assertTrue(data['fichier_url'])


class CloudinaryConfigurationTests(SimpleTestCase):
    def test_cloudinary_settings_support_cloudinary_url(self):
        with mock.patch.dict(
            os.environ,
            {'CLOUDINARY_URL': 'cloudinary://123456:secret@demo-cloud'},
            clear=False,
        ):
            for env_name in (
                'CLOUDINARY_CLOUD_NAME',
                'CLOUDINARY_API_KEY',
                'CLOUDINARY_API_SECRET',
            ):
                os.environ.pop(env_name, None)

            reloaded_settings = importlib.reload(settings_module)

            self.assertTrue(reloaded_settings._cloudinary_ready)
            self.assertEqual(reloaded_settings.CLOUDINARY_STORAGE['CLOUD_NAME'], 'demo-cloud')
            self.assertEqual(reloaded_settings.CLOUDINARY_STORAGE['API_KEY'], '123456')
            self.assertEqual(reloaded_settings.CLOUDINARY_STORAGE['API_SECRET'], 'secret')


class PublicIncidentAttachmentTests(TestCase):
    def test_public_incident_accepts_allowed_attachments(self):
        province = Province.objects.create(nom='Kinshasa', code='KIN')
        Employeur.objects.create(
            nom='Entreprise privée',
            secteur='services',
            province=province,
        )

        payload = {
            'employeur': 'Entreprise privée',
            'employeur_address': '123 Rue',
            'ville': 'Kinshasa',
            'province': province.id,
            'type_incident': 'salaire',
            'description': 'Description suffisante pour le test de pièces jointes.',
            'le_fautif': 'Entreprise privée',
            'est_anonyme': True,
            'email_contact_anonyme': 'test@example.com',
            'telephone_contact_anonyme': '',
            'secteur': 'services',
            'autre_secteur': '',
            'confirm_anonymous': True,
        }
        response = self.client.post(
            reverse('api:public_incident_create'),
            payload,
            format='multipart',
        )

        self.assertEqual(response.status_code, 201)

    def test_public_incident_accepts_non_existing_province_name(self):
        payload = {
            'employeur': 'Entreprise privée',
            'employeur_address': '123 Rue',
            'ville': 'Kinshasa',
            'province': 'Bas-Uélé',
            'type_incident': 'salaire',
            'description': 'Description suffisante pour valider le choix d’une province non encore créée.',
            'le_fautif': 'Entreprise privée',
            'est_anonyme': True,
            'email_contact_anonyme': 'test@example.com',
            'telephone_contact_anonyme': '',
            'secteur': 'services',
            'autre_secteur': '',
            'confirm_anonymous': True,
        }

        response = self.client.post(
            reverse('api:public_incident_create'),
            payload,
            format='multipart',
        )

        self.assertEqual(response.status_code, 201)
        self.assertTrue(Province.objects.filter(nom__iexact='Bas-Uélé').exists())