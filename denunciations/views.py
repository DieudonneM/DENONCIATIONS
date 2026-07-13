"""Vues pour l'application denunciations."""

from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse
from django.db.models import Q
from django.contrib import messages

from .models import Incident, Commentaire
from .forms import FilterIncidentForm
from core.utils import get_incidents_by_user_role, check_user_can_view_incident
from core.views import (
    IncidentPublicFormView,
    IncidentSuccessView,
    SearchIncidentView,
    IncidentDetailView,
    UpdateIncidentStatusView,
    AssignIncidentView,
)
from users.auth_backends import user_is_agent, user_is_admin
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


class ToggleCommentVisibility(LoginRequiredMixin, View):
    """Basculer la visibilité d'un commentaire (public <-> interne)."""
    login_url = 'users:login'

    def post(self, request, comment_id):
        try:
            commentaire = Commentaire.objects.get(id=comment_id)
        except Commentaire.DoesNotExist:
            return JsonResponse({'error': 'Commentaire introuvable'}, status=404)

        # Autorisation : agents et admins seulement
        if not (user_is_agent(request.user) or user_is_admin(request.user)):
            return JsonResponse({'error': 'Permission refusée'}, status=403)

        commentaire.type_commentaire = 'public' if commentaire.type_commentaire == 'interne' else 'interne'
        commentaire.save()

        messages.success(request, 'Visibilité du commentaire modifiée.')
        return redirect(request.META.get('HTTP_REFERER', '/'))


class ExportIncidentsExcel(LoginRequiredMixin, View):
    """Exporter une liste d'incidents (filtrée) au format Excel."""
    login_url = 'users:login'

    def get(self, request):
        # Filtrer selon paramètres GET simples
        qs = get_incidents_by_user_role(request.user).select_related('employeur', 'province')
        statut = request.GET.get('statut')
        ttype = request.GET.get('type_incident')
        search = request.GET.get('search')
        if statut:
            qs = qs.filter(statut=statut)
        if ttype:
            qs = qs.filter(type_incident=ttype)
        if search:
            qs = qs.filter(
                Q(code_suivi__icontains=search) |
                Q(employeur__nom__icontains=search) |
                Q(ville__icontains=search)
            )

        # Créer workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Incidents'

        headers = ['Code', 'Type', 'Employeur', 'Ville', 'Province', 'Statut', 'Date création', 'Description', 'Le fautif']
        ws.append(headers)

        for inc in qs.order_by('-date_creation'):
            ws.append([
                inc.code_suivi,
                inc.get_type_incident_display(),
                inc.employeur.nom if inc.employeur else '',
                inc.ville,
                inc.province.nom if inc.province else '',
                inc.get_statut_display(),
                inc.date_creation.strftime('%Y-%m-%d %H:%M'),
                inc.description,
                getattr(inc, 'le_fautif', '') or ''
            ])

        # Ajuster largeur colonnes
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(i)].width = min(50, max_length + 2)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=incidents_export.xlsx'
        wb.save(response)
        return response


class IncidentsListView(LoginRequiredMixin, View):
    """Liste paginée et filtrable des incidents (exportable)."""
    login_url = 'users:login'

    def get(self, request):
        user = request.user
        qs = get_incidents_by_user_role(user).select_related('employeur', 'province')

        # Appliquer filtres
        form = FilterIncidentForm(request.GET or None)
        if form.is_valid():
            statut = form.cleaned_data.get('statut')
            ttype = form.cleaned_data.get('type_incident')
            search = form.cleaned_data.get('search')

            if statut:
                qs = qs.filter(statut=statut)
            if ttype:
                qs = qs.filter(type_incident=ttype)
            if search:
                qs = qs.filter(
                    Q(code_suivi__icontains=search) |
                    Q(employeur__nom__icontains=search) |
                    Q(ville__icontains=search)
                )

        # Tri par date
        qs = qs.order_by('-date_creation')

        # Pagination (15 par page)
        paginator = Paginator(qs, 15)
        page = request.GET.get('page', 1)
        try:
            incidents_page = paginator.page(page)
        except PageNotAnInteger:
            incidents_page = paginator.page(1)
        except EmptyPage:
            incidents_page = paginator.page(paginator.num_pages)

        # Prepare context
        context = {
            'page_title': 'Liste des Dénonciations',
            'filter_form': form,
            'incidents_page': incidents_page,
            'paginator': paginator,
            'export_endpoint': f"{request.scheme}://{request.get_host()}{request.path.replace('incidents/', 'export/xlsx/')}",
        }
        # If AJAX request, return only the fragment HTML
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('ajax'):
            html = render(request, 'denunciations/_incidents_fragment.html', context).content.decode('utf-8')
            return JsonResponse({'html': html})

        return render(request, 'denunciations/incidents_list.html', context)


class ExportIncidentExcel(LoginRequiredMixin, View):
    """Exporter un incident unique (détails + commentaires) en Excel."""
    login_url = 'users:login'

    def get(self, request, code):
        incident = get_object_or_404(Incident, code_suivi=code)

        # Vérifier permission de visualiser
        if not check_user_can_view_incident(request.user, incident):
            return render(request, 'core/error_403.html', status=403)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Incident'

        # Détails incident
        rows = [
            ('Code', incident.code_suivi),
            ('Type', incident.get_type_incident_display()),
            ('Employeur', incident.employeur.nom if incident.employeur else ''),
            ('Ville', incident.ville),
            ('Province', incident.province.nom if incident.province else ''),
            ('Statut', incident.get_statut_display()),
            ('Date création', incident.date_creation.strftime('%Y-%m-%d %H:%M')),
            ('Description', incident.description),
            ('Le fautif', getattr(incident, 'le_fautif', '') or ''),
            ('Pièces jointes', ', '.join([p.fichier.name for p in incident.pieces_jointes.all()])),
        ]

        for key, val in rows:
            ws.append([key, val])

        # Comments sheet
        ws2 = wb.create_sheet(title='Commentaires')
        ws2.append(['Auteur', 'Type', 'Date', 'Texte'])
        for c in incident.commentaires.all().order_by('date_creation'):
            ws2.append([str(c.auteur) if c.auteur else 'Anonyme', c.type_commentaire, c.date_creation.strftime('%Y-%m-%d %H:%M'), c.texte])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=incident_{incident.code_suivi}.xlsx'
        wb.save(response)
        return response
